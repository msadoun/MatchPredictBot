"""Merge-only import from admin Excel exports — never deletes existing predictions."""

import logging
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook

from config import ALKORAM3NA_GROUP_CHAT_ID, DATABASE_PATH, KM3NA_GROUP_CHAT_ID
from database import (
    bulk_set_group_manual_points,
    get_db,
    get_match,
    get_match_by_teams,
    merge_prediction_if_missing,
    recalculate_all_prediction_points,
    resolve_user_ref,
    upsert_user,
    _row_to_user,
)
from group_standings import (
    ALKORAM3NA_GROUP_USERNAME,
    KM3NA_GROUP_USERNAME,
    PREDEFINED_GROUP_STANDINGS,
)
from prediction_reports import EXCEL_HEADERS, EXPORTS_DIR

logger = logging.getLogger(__name__)

IMPORTS_DIR = DATABASE_PATH.parent / "imports"
PREDICTION_RE = re.compile(r"^(.+?) (\d+)-(\d+) (.+)$")


@dataclass
class ExcelImportResult:
    merged: int = 0
    skipped: int = 0
    points_updated: int = 0
    users_created: int = 0
    users_not_found: list[str] | None = None
    matches_not_found: list[str] | None = None
    group_points_applied: int = 0
    group_points_missing: int = 0
    files_read: list[str] | None = None

    def __post_init__(self) -> None:
        if self.users_not_found is None:
            self.users_not_found = []
        if self.matches_not_found is None:
            self.matches_not_found = []
        if self.files_read is None:
            self.files_read = []


def _parse_points(value: object) -> int | None:
    if value is None or value == "":
        return None
    try:
        return int(value)
    except (TypeError, ValueError):
        return None


def _parse_prediction(text: str) -> tuple[str, int, int, str] | None:
    cleaned = text.strip()
    if not cleaned:
        return None
    match = PREDICTION_RE.match(cleaned)
    if not match:
        return None
    return match.group(1).strip(), int(match.group(2)), int(match.group(3)), match.group(4).strip()


def _find_match(home_team: str, away_team: str):
    exact = get_match_by_teams(home_team, away_team)
    if exact:
        return exact
    with get_db() as conn:
        rows = conn.execute(
            "SELECT id, home_team, away_team FROM matches"
        ).fetchall()
    for row in rows:
        if row["home_team"].strip() == home_team and row["away_team"].strip() == away_team:
            return get_match(int(row["id"]))
    return None


def _ensure_user(name: str) -> tuple[object | None, bool]:
    user = resolve_user_ref(name)
    if user:
        return user, False

    from config import M2USAB_TELEGRAM_ID, M2USAB_USERNAME

    if name.strip().lstrip("@").lower() == M2USAB_USERNAME:
        user = resolve_user_ref("M2usab")
        if user:
            return user, False
        return upsert_user(M2USAB_TELEGRAM_ID, M2USAB_USERNAME, "M2usab"), True

    with get_db() as conn:
        row = conn.execute(
            """
            SELECT * FROM users
            WHERE LOWER(display_name) = LOWER(?)
               OR LOWER(username) = LOWER(?)
            LIMIT 1
            """,
            (name.strip(), name.strip().lstrip("@")),
        ).fetchone()
    if row:
        return _row_to_user(row), False
    return None, False


def apply_predefined_group_standings() -> tuple[int, int]:
    """Set manual base points from predefined lists — does not touch predictions."""
    applied_total = 0
    missing_total = 0
    groups = [
        (ALKORAM3NA_GROUP_CHAT_ID, ALKORAM3NA_GROUP_USERNAME),
        (KM3NA_GROUP_CHAT_ID, KM3NA_GROUP_USERNAME),
    ]
    for chat_id_raw, group_key in groups:
        if not chat_id_raw.lstrip("-").isdigit():
            continue
        standings = PREDEFINED_GROUP_STANDINGS.get(group_key, [])
        if not standings:
            continue
        applied, not_found = bulk_set_group_manual_points(int(chat_id_raw), standings)
        applied_total += len(applied)
        missing_total += len(not_found)
    return applied_total, missing_total


def import_predictions_from_excel(path: Path) -> ExcelImportResult:
    """Merge rows from an export-format .xlsx file. Never deletes existing data."""
    result = ExcelImportResult()
    if not path.is_file():
        return result

    workbook = load_workbook(path, read_only=True, data_only=True)
    result.files_read.append(path.name)

    for sheet in workbook.worksheets:
        rows = sheet.iter_rows(values_only=True)
        header: tuple[str, ...] | None = None
        for row in rows:
            if not row or not any(cell is not None and str(cell).strip() for cell in row):
                continue
            cells = tuple("" if cell is None else str(cell).strip() for cell in row)
            if header is None:
                lowered = [cell.lower() for cell in cells[:5]]
                if lowered[:1] == [EXCEL_HEADERS[0]] or "user" in lowered[0]:
                    header = cells
                    continue
                header = EXCEL_HEADERS

            if len(cells) < 2:
                continue

            user_name = cells[0].strip()
            prediction_text = cells[1].strip() if len(cells) > 1 else ""
            points_value = _parse_points(cells[4] if len(cells) > 4 else None)

            if not user_name:
                continue
            if not prediction_text:
                continue

            parsed = _parse_prediction(prediction_text)
            if not parsed:
                result.skipped += 1
                continue

            home_team, home_score, away_score, away_team = parsed
            match = _find_match(home_team, away_team)
            if not match:
                label = f"{home_team} vs {away_team}"
                if label not in result.matches_not_found:
                    result.matches_not_found.append(label)
                result.skipped += 1
                continue

            user, created = _ensure_user(user_name)
            if not user:
                if user_name not in result.users_not_found:
                    result.users_not_found.append(user_name)
                result.skipped += 1
                continue
            if created:
                result.users_created += 1

            action = merge_prediction_if_missing(
                user.id,
                match.id,
                home_score,
                away_score,
                points=points_value,
            )
            if action == "merged":
                result.merged += 1
            elif action == "points_updated":
                result.points_updated += 1
            else:
                result.skipped += 1

    workbook.close()
    return result


def import_all_excel_sources(*, include_exports: bool = True) -> ExcelImportResult:
    """Import every .xlsx in data/imports and optionally data/exports (merge only)."""
    combined = ExcelImportResult()
    paths: list[Path] = []

    if IMPORTS_DIR.is_dir():
        paths.extend(sorted(IMPORTS_DIR.glob("*.xlsx")))
    if include_exports and EXPORTS_DIR.is_dir():
        paths.extend(sorted(EXPORTS_DIR.glob("*.xlsx"), key=lambda p: p.stat().st_mtime))

    seen: set[str] = set()
    for path in paths:
        if path.name in seen:
            continue
        seen.add(path.name)
        partial = import_predictions_from_excel(path)
        combined.merged += partial.merged
        combined.skipped += partial.skipped
        combined.points_updated += partial.points_updated
        combined.users_created += partial.users_created
        combined.files_read.extend(partial.files_read)
        for item in partial.users_not_found:
            if item not in combined.users_not_found:
                combined.users_not_found.append(item)
        for item in partial.matches_not_found:
            if item not in combined.matches_not_found:
                combined.matches_not_found.append(item)

    applied, missing = apply_predefined_group_standings()
    combined.group_points_applied = applied
    combined.group_points_missing = missing
    recalculate_all_prediction_points()
    return combined


def import_if_database_empty() -> ExcelImportResult | None:
    from prediction_backup import count_predictions

    if count_predictions() > 0:
        return None
    result = import_all_excel_sources()
    if result.merged or result.points_updated or result.group_points_applied:
        logger.info(
            "Excel auto-import: merged=%d points_updated=%d group_pts=%d",
            result.merged,
            result.points_updated,
            result.group_points_applied,
        )
        return result
    return None
