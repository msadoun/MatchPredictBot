import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from config import DATABASE_PATH
from database import Match, User, get_db, list_matches
from worldcup2026 import (
    GROUP_STAGE_LABEL,
    is_group_stage_label,
    match_day_date,
    match_on_day,
    ordered_match_days,
    ordered_report_stages,
    stage_from_kickoff,
)

EXCEL_HEADERS = ("user name", "prediction", "result", "round", "points", "double")


@dataclass
class PredictionCell:
    home_score: int
    away_score: int
    points: int | None
    is_doubled: bool = False


@dataclass
class PredictionReport:
    scope_type: str
    scope_key: str
    scope_label: str
    matches: list[Match]
    users: list[User]
    predictions: dict[tuple[int, int], PredictionCell]
    always_include_user_ids: frozenset[int] = frozenset()
    extra_display_names: tuple[str, ...] = ()


@dataclass
class SavedExport:
    id: int
    scope_type: str
    scope_key: str
    scope_label: str
    file_path: str
    match_count: int
    user_count: int
    prediction_count: int
    saved_at: str


EXPORTS_DIR = DATABASE_PATH.parent / "exports"
# Excel exports are append-only: new files only, existing exports are never deleted.


def _safe_filename_part(text: str) -> str:
    cleaned = re.sub(r"[^\w\-]+", "_", text, flags=re.UNICODE).strip("_")
    return cleaned or "report"


def _match_round(match: Match) -> str:
    if match.kickoff_at:
        stage = stage_from_kickoff(match.kickoff_at)
        if stage:
            return stage
    return ""


def _score_line(match: Match, home_score: int, away_score: int) -> str:
    return f"{match.home_team} {home_score}-{away_score} {match.away_team}"


def _actual_result(match: Match) -> str:
    if match.home_score is None or match.away_score is None:
        return ""
    return _score_line(match, match.home_score, match.away_score)


def matches_for_scope(scope_type: str, scope_key: str) -> list[Match]:
    all_matches = list_matches(open_only=False)
    if scope_type == "day":
        return [
            m
            for m in all_matches
            if m.kickoff_at and match_on_day(m.kickoff_at, scope_key)
        ]
    if scope_type == "stage":
        return [
            m
            for m in all_matches
            if m.kickoff_at and stage_from_kickoff(m.kickoff_at) == scope_key
        ]
    if scope_type == "group_stage":
        return [
            m
            for m in all_matches
            if m.kickoff_at and is_group_stage_label(stage_from_kickoff(m.kickoff_at) or "")
        ]
    return []


def scope_label(scope_type: str, scope_key: str) -> str:
    if scope_type == "day":
        return f"يوم {scope_key}"
    if scope_type == "group_stage":
        return GROUP_STAGE_LABEL
    return scope_key


def _resolve_excel_roster_user(ref: str):
    from group_standings import resolve_excel_user

    return resolve_excel_user(ref)


def _resolve_always_include_users(
    existing_users: list[User],
) -> tuple[list[User], frozenset[int], tuple[str, ...]]:
    from group_standings import EXCEL_ALWAYS_INCLUDE_USERS

    by_id = {user.id: user for user in existing_users}
    always_ids: set[int] = set()
    extra_names: list[str] = []

    for ref in EXCEL_ALWAYS_INCLUDE_USERS:
        user = _resolve_excel_roster_user(ref)
        if user:
            always_ids.add(user.id)
            if user.id not in by_id:
                by_id[user.id] = user
            continue
        label = ref.strip().lstrip("@")
        if label and label not in extra_names:
            extra_names.append(label)

    merged = sorted(by_id.values(), key=lambda user: user.display_name.lower())
    return merged, frozenset(always_ids), tuple(extra_names)


def build_prediction_report(scope_type: str, scope_key: str) -> PredictionReport:
    matches = matches_for_scope(scope_type, scope_key)
    match_ids = [m.id for m in matches]
    users: list[User] = []
    predictions: dict[tuple[int, int], PredictionCell] = {}

    if match_ids:
        placeholders = ",".join("?" for _ in match_ids)
        with get_db() as conn:
            user_rows = conn.execute(
                f"""
                SELECT DISTINCT u.*
                FROM users u
                INNER JOIN predictions p ON p.user_id = u.id
                WHERE p.match_id IN ({placeholders})
                ORDER BY u.display_name ASC
                """,
                match_ids,
            ).fetchall()
            pred_rows = conn.execute(
                f"""
                SELECT p.user_id, p.match_id, p.home_score, p.away_score, p.points,
                       p.is_doubled
                FROM predictions p
                WHERE p.match_id IN ({placeholders})
                """,
                match_ids,
            ).fetchall()

        from database import _row_to_user

        users = [_row_to_user(row) for row in user_rows]
        for row in pred_rows:
            predictions[(row["user_id"], row["match_id"])] = PredictionCell(
                home_score=row["home_score"],
                away_score=row["away_score"],
                points=row["points"],
                is_doubled=bool(row["is_doubled"] or 0),
            )

    users, always_include_user_ids, extra_display_names = _resolve_always_include_users(
        users
    )

    return PredictionReport(
        scope_type=scope_type,
        scope_key=scope_key,
        scope_label=scope_label(scope_type, scope_key),
        matches=matches,
        users=users,
        predictions=predictions,
        always_include_user_ids=always_include_user_ids,
        extra_display_names=extra_display_names,
    )


def report_summary_text(report: PredictionReport, *, max_users: int = 12) -> str:
    total_preds = len(report.predictions)
    participant_count = len(report.users) + len(report.extra_display_names)
    lines = [
        f"📊 {report.scope_label}",
        f"• {len(report.matches)} مباراة",
        f"• {participant_count} لاعب",
        f"• {total_preds} توقع",
    ]
    if not report.matches:
        lines.append("\nلا توجد مباريات في هذا النطاق.")
        return "\n".join(lines)

    lines.append("\n👥 ملخص سريع:")
    shown = 0
    for user in report.users:
        if shown >= max_users:
            lines.append(f"... و {participant_count - max_users} لاعب آخر")
            break
        picks: list[str] = []
        for match in report.matches:
            cell = report.predictions.get((user.id, match.id))
            if cell:
                picks.append(f"#{match.id} {cell.home_score}-{cell.away_score}")
        name = user.display_name
        if user.username:
            name += f" (@{user.username})"
        if picks:
            lines.append(f"• {name}: {', '.join(picks)}")
        else:
            lines.append(f"• {name}: —")
        shown += 1

    for name in report.extra_display_names:
        if shown >= max_users:
            lines.append(f"... و {participant_count - max_users} لاعب آخر")
            break
        lines.append(f"• {name}: —")
        shown += 1

    if participant_count == 0:
        lines.append("لا توجد توقعات بعد في هذا النطاق.")
    return "\n".join(lines)


def _double_points_label(is_doubled: bool) -> str:
    return "yes" if is_doubled else "no"


MATCH_PHOTO_HEADERS = ("user name", "prediction", "double")
MATCH_PHOTO_FILL = "C5D9F1"
_LTR = "\u200e"


def _double_points_table_label(is_doubled: bool) -> str:
    return "X2" if is_doubled else ""


def build_match_table_rows(match: Match) -> list[tuple[str, str, bool]]:
    from database import list_predictions_for_match

    rows: list[tuple[str, str, bool]] = []
    for user, prediction in list_predictions_for_match(match.id):
        rows.append(
            (
                user.display_name,
                _score_line(match, prediction.home_score, prediction.away_score),
                bool(prediction.is_doubled),
            )
        )
    return rows


def build_match_photo_rows(match: Match) -> list[tuple[str, str, str]]:
    return [
        (user, prediction, "yes" if is_doubled else "no")
        for user, prediction, is_doubled in build_match_table_rows(match)
    ]


def format_match_prediction_table(match: Match) -> str:
    rows = build_match_table_rows(match)
    lines = [
        f"📋 توقعات المباراة #{match.id}",
        f"{match.home_team} ضد {match.away_team}",
    ]
    if match.kickoff_at:
        lines.append(f"الموعد: {match.kickoff_at}")
    if match.home_score is not None and match.away_score is not None:
        lines.append(f"النتيجة: {match.home_score}-{match.away_score}")
    else:
        lines.append("النتيجة: لم تُسجَّل بعد")

    if not rows:
        lines.append("")
        lines.append("لا توجد توقعات لهذه المباراة بعد.")
        return "\n".join(lines)

    user_blocks: list[str] = []
    for index, (user, prediction, is_doubled) in enumerate(rows, start=1):
        block_lines = [
            f"{_LTR}{index}. {user}",
            f"{_LTR}   التوقع: {prediction}",
        ]
        double_label = _double_points_table_label(is_doubled)
        if double_label:
            block_lines.append(f"{_LTR}   {double_label}")
        user_blocks.append("\n".join(block_lines))

    lines.extend(["", "━━━━━━━━━━━━━━", ""])
    lines.append("\n\n".join(user_blocks))
    lines.extend(["", "━━━━━━━━━━━━━━", "", f"عدد التوقعات: {len(rows)}"])
    return "\n".join(lines)


def build_match_photo_workbook(match: Match) -> Workbook:
    from openpyxl.styles import Alignment, Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = f"match_{match.id}"[:31]

    header_fill = PatternFill(fill_type="solid", fgColor=MATCH_PHOTO_FILL)
    alt_fill = PatternFill(fill_type="solid", fgColor=MATCH_PHOTO_FILL)
    white_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")

    sheet.append(list(MATCH_PHOTO_HEADERS))
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="right")

    for index, row in enumerate(build_match_photo_rows(match)):
        sheet.append(list(row))
        row_fill = white_fill if index % 2 == 0 else alt_fill
        for cell in sheet[index + 2]:
            cell.fill = row_fill
            cell.alignment = Alignment(horizontal="right")

    for column_index, header in enumerate(MATCH_PHOTO_HEADERS, start=1):
        column = get_column_letter(column_index)
        max_len = len(header)
        for row in sheet.iter_rows(
            min_row=2,
            min_col=column_index,
            max_col=column_index,
            values_only=True,
        ):
            if row[0] is not None:
                max_len = max(max_len, len(str(row[0])))
        sheet.column_dimensions[column].width = min(max_len + 2, 50)

    return workbook


def _excel_rows(
    report: PredictionReport,
) -> list[tuple[str, str, str, str, int | str, str]]:
    rows: list[tuple[str, str, str, str, int | str, str]] = []
    for user in report.users:
        for match in report.matches:
            cell = report.predictions.get((user.id, match.id))
            if cell:
                points = cell.points if cell.points is not None else 0
                rows.append(
                    (
                        user.display_name,
                        _score_line(match, cell.home_score, cell.away_score),
                        _actual_result(match),
                        _match_round(match),
                        points,
                        _double_points_label(cell.is_doubled),
                    )
                )
            elif user.id in report.always_include_user_ids:
                rows.append(
                    (
                        user.display_name,
                        "",
                        _actual_result(match),
                        _match_round(match),
                        "",
                        "",
                    )
                )

    for name in report.extra_display_names:
        for match in report.matches:
            rows.append(
                (
                    name,
                    "",
                    _actual_result(match),
                    _match_round(match),
                    "",
                    "",
                )
            )
    return rows


def report_to_excel(report: PredictionReport) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = report.scope_label[:31] or "predictions"

    sheet.append(list(EXCEL_HEADERS))
    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row in _excel_rows(report):
        sheet.append(list(row))

    for column_index, header in enumerate(EXCEL_HEADERS, start=1):
        column = get_column_letter(column_index)
        max_len = len(header)
        for row in sheet.iter_rows(
            min_row=2,
            min_col=column_index,
            max_col=column_index,
            values_only=True,
        ):
            if row[0] is not None:
                max_len = max(max_len, len(str(row[0])))
        sheet.column_dimensions[column].width = min(max_len + 2, 50)

    return workbook


def save_prediction_export(
    report: PredictionReport,
    *,
    saved_by_telegram_id: int | None = None,
) -> tuple[Path, SavedExport]:
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    filename = (
        f"{report.scope_type}_{_safe_filename_part(report.scope_key)}_{stamp}.xlsx"
    )
    path = EXPORTS_DIR / filename
    report_to_excel(report).save(path)

    saved_at = datetime.utcnow().isoformat()
    with get_db() as conn:
        row = conn.execute(
            """
            INSERT INTO prediction_exports (
                scope_type, scope_key, scope_label, file_path,
                match_count, user_count, prediction_count,
                saved_at, saved_by_telegram_id
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                report.scope_type,
                report.scope_key,
                report.scope_label,
                str(path),
                len(report.matches),
                len(report.users) + len(report.extra_display_names),
                len(report.predictions),
                saved_at,
                saved_by_telegram_id,
            ),
        )
        export_id = row.lastrowid

    saved = SavedExport(
        id=int(export_id),
        scope_type=report.scope_type,
        scope_key=report.scope_key,
        scope_label=report.scope_label,
        file_path=str(path),
        match_count=len(report.matches),
        user_count=len(report.users) + len(report.extra_display_names),
        prediction_count=len(report.predictions),
        saved_at=saved_at,
    )
    return path, saved


def list_saved_exports(limit: int = 20) -> list[SavedExport]:
    with get_db() as conn:
        rows = conn.execute(
            """
            SELECT * FROM prediction_exports
            ORDER BY saved_at DESC
            LIMIT ?
            """,
            (limit,),
        ).fetchall()
    return [
        SavedExport(
            id=row["id"],
            scope_type=row["scope_type"],
            scope_key=row["scope_key"],
            scope_label=row["scope_label"],
            file_path=row["file_path"],
            match_count=row["match_count"],
            user_count=row["user_count"],
            prediction_count=row["prediction_count"],
            saved_at=row["saved_at"],
        )
        for row in rows
    ]


def get_saved_export(export_id: int) -> SavedExport | None:
    with get_db() as conn:
        row = conn.execute(
            "SELECT * FROM prediction_exports WHERE id = ?",
            (export_id,),
        ).fetchone()
    if not row:
        return None
    return SavedExport(
        id=row["id"],
        scope_type=row["scope_type"],
        scope_key=row["scope_key"],
        scope_label=row["scope_label"],
        file_path=row["file_path"],
        match_count=row["match_count"],
        user_count=row["user_count"],
        prediction_count=row["prediction_count"],
        saved_at=row["saved_at"],
    )


def list_available_days() -> list[str]:
    db_days = {
        match_day_date(m.kickoff_at)
        for m in list_matches(open_only=False)
        if m.kickoff_at
    }
    ordered = [d for d in ordered_match_days() if d in db_days]
    for day in sorted(db_days):
        if day not in ordered:
            ordered.append(day)
    return ordered


def list_available_stages() -> list[str]:
    db_stages = {
        stage_from_kickoff(m.kickoff_at)
        for m in list_matches(open_only=False)
        if m.kickoff_at and stage_from_kickoff(m.kickoff_at)
    }
    ordered = [s for s in ordered_report_stages() if s in db_stages]
    for stage in sorted(db_stages):
        if stage not in ordered:
            ordered.append(stage)
    return ordered
